"""
SIMPLEST migration path — one Excel file per SUK (the whole spreadsheet,
every tab included), no Apps Script involved at all.

HOW TO EXPORT (do this once per SUK, in your browser):
    Open the SUK's Google Sheet → File → Download → Microsoft Excel (.xlsx)

SAVE each downloaded file into one folder, renamed to exactly the suk_key:
    bannerghatta.xlsx
    peenya-2nd-stage.xlsx
    banashankari.xlsx
    marathahalli.xlsx
    electronic-city.xlsx
    garvebhavi-palya.xlsx

You don't need to touch Bookings/Bhadra/Matri/Savan/Satsang tabs individually —
this script reads whichever of those tabs exist inside each workbook automatically.
"Photos" / "Photos-copy" tabs are ignored here (photos migrate separately later).

USAGE:
    # Preview only, writes nothing:
    python import_from_xlsx.py --dir xlsx_export --dry-run

    # Actually import:
    python import_from_xlsx.py --dir xlsx_export

Safe to re-run: for each suk_key + sheet type, existing rows are deleted and
replaced fresh, so re-running never creates duplicates.
"""
import sys
import asyncio
import argparse
import pathlib

sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))

from openpyxl import load_workbook
from sqlalchemy import delete
from backend.shared.db import get_session, init_db, db_configured
from backend.shared.db_models import Booking, EventBooking

EVENT_SHEETS = {"Satsang": "satsang", "Bhadra": "bhadra", "Matri": "matri", "Savan": "savan"}
ALL_SHEET_NAMES = ["Bookings", "Satsang", "Bhadra", "Matri", "Savan"]

VALID_SUKS = {
    "bannerghatta", "peenya-2nd-stage", "banashankari",
    "marathahalli", "electronic-city", "garvebhavi-palya",
}


def normalize_date(val) -> str:
    """Handles both plain strings and real datetime objects (openpyxl
    returns real Python datetimes for date-formatted cells)."""
    if val is None or val == "":
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    if len(val) >= 10 and val[4] == "-" and val[7] == "-":
        return val[:10]
    return val


def cell_str(val) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        # Excel/Sheets often stores "number-looking" columns (like mobile
        # numbers) as floats. str(7798702822.0) == "7798702822.0" — strip
        # the trailing .0 for whole numbers so mobile numbers stay clean.
        if val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()


def read_sheet_rows(ws) -> list[list]:
    """Skip header row (row 1) and any fully-blank rows."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            rows.append(list(row))
    return rows


async def import_bookings_sheet(suk_key: str, ws, dry_run: bool) -> int:
    rows = read_sheet_rows(ws)
    print(f"    Bookings   → {len(rows)} row(s)" + (" (dry run)" if dry_run else ""))
    if dry_run or not rows:
        return len(rows)

    async with get_session() as session:
        await session.execute(delete(Booking).where(Booking.suk_key == suk_key))
        for r in rows:
            r = r + [None] * (9 - len(r))
            session.add(Booking(
                suk_key=suk_key,
                name=cell_str(r[4]), mobile=cell_str(r[5]), place=cell_str(r[6]),
                maps_link="",  # already embedded inside `place` text for this sheet type
                date=normalize_date(r[1]), time=cell_str(r[3]), day=cell_str(r[2]),
            ))
    return len(rows)


async def import_event_sheet(suk_key: str, event_type: str, sheet_name: str, ws, dry_run: bool) -> int:
    rows = read_sheet_rows(ws)
    print(f"    {sheet_name:10s} → {len(rows)} row(s)" + (" (dry run)" if dry_run else ""))
    if dry_run or not rows:
        return len(rows)

    async with get_session() as session:
        await session.execute(delete(EventBooking).where(
            EventBooking.suk_key == suk_key, EventBooking.event_type == event_type
        ))
        for r in rows:
            r = r + [None] * (10 - len(r))
            session.add(EventBooking(
                suk_key=suk_key, event_type=event_type,
                name=cell_str(r[4]), mobile=cell_str(r[5]), venue=cell_str(r[6]),
                maps_link=cell_str(r[7]), date=normalize_date(r[1]), time=cell_str(r[3]),
                hosted_by=cell_str(r[8]), occasion="",
            ))
    return len(rows)


async def main(dir_path: str, dry_run: bool):
    if not dry_run and not db_configured():
        print("❌ DATABASE_URL is not set.")
        return
    if not dry_run:
        await init_db()

    folder = pathlib.Path(dir_path)
    if not folder.is_dir():
        print(f"❌ Folder not found: {folder}")
        return

    xlsx_files = sorted(folder.glob("*.xlsx"))
    if not xlsx_files:
        print(f"❌ No .xlsx files found in {folder}")
        return

    grand_total = 0
    for f in xlsx_files:
        suk_key = f.stem
        if suk_key not in VALID_SUKS:
            print(f"⚠️  Skipping {f.name} — '{suk_key}' isn't a recognized suk_key "
                  f"(rename the file to match one of: {', '.join(sorted(VALID_SUKS))})")
            continue

        print(f"\n📍 {suk_key}  ({f.name})")
        wb = load_workbook(f, data_only=True, read_only=True)
        suk_total = 0

        for sheet_name in ALL_SHEET_NAMES:
            if sheet_name not in wb.sheetnames:
                print(f"    {sheet_name:10s} → (tab not present, skipped)")
                continue
            ws = wb[sheet_name]
            if sheet_name == "Bookings":
                suk_total += await import_bookings_sheet(suk_key, ws, dry_run)
            else:
                suk_total += await import_event_sheet(suk_key, EVENT_SHEETS[sheet_name], sheet_name, ws, dry_run)

        wb.close()
        print(f"    ── {suk_key} total: {suk_total} row(s)")
        grand_total += suk_total

    print(f"\n{'Would import' if dry_run else 'Imported'} {grand_total} row(s) total.")
    if dry_run:
        print("This was a DRY RUN — nothing was written. Re-run without --dry-run to actually import.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", default="xlsx_export", help="Folder containing the exported .xlsx files")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    asyncio.run(main(args.dir, args.dry_run))